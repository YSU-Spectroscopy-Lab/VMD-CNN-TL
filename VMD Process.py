from vmdpy import VMD
import numpy as np
import pandas as pd
from openpyxl import Workbook
import os


def read_data(path):
    df = pd.read_excel(path, header=None)
    data = np.array(df)
    data_x = data[:, 0:-2]
    data_y = data[:, -2:]
    return data_x, data_y


def run_VMD(data, num):
    data = data
    # some sample parameters for VMD
    alpha = 2000  # moderate bandwidth constraint
    tau = 0.  # noise-tolerance (no strict fidelity enforcement)
    K = num  # mode number
    DC = 0  # no DC part imposed
    init = 0  # initialize omegas uniformly
    tol = 1e-7
    uu_ = []
    for line in data:
        u, u_hat, omega = VMD(line, alpha, tau, K, DC, init, tol)
        uu_.append(u)
    print("VMD processing completed！")
    return u, u_hat, omega, uu_


def writeinexcel(l, y, name, aaa):
    wb = Workbook()  # 新建一个excel文件
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    data_x = l
    data_y = y
    n_ = True
    for i in range(len(data_x)):
        data = 0
        for j in range(len(data_x[0][0])):
            for ii in range(aaa - 1):
                data += data_x[i][ii][j]
            if n_:
                ws.cell(1, j + 1, j + 1)
            ws.cell(i + 2, j + 1, data)
        n_ = False
        ws.cell(i + 2, len(data_x[0][0]) + 1, data_y[i][0])
        ws.cell(i + 2, len(data_x[0][0]) + 2, data_y[i][1])
        print("###################Completed No." + str(i + 1) + "/" + str(len(data_x)) + " line#####################")
    ws.cell(1, len(data_x[0][0]) + 1, len(data_x[0][0]) + 1)
    ws.cell(1, len(data_x[0][0]) + 2, len(data_x[0][0]) + 2)
    wb.save(name)
    print("Program operation completed!")


def del_files(name):
    if os.path.exists(name):
        os.remove(name)
    print("Data environment cleanup succeeded!")


if __name__ == '__main__':
    '''    
    When the final two documents (Combined_gas_data.xlsx and Mixed_gas_data.xlsx) have been generated:
        Insert a line in the first row of each file, from 1-424 (e.g. the format of the first row in the file Spectral Data/Test data/VMD-Test.xlsx)
    '''
    is_single_gas = False
    '''When is_single_gas = True, it takes a few minutes to run, depending on the configuration of the computer.'''
    if is_single_gas:
        path = "Spectral Data/Original spectral data/TL_Data1/Combined_gas_data.xlsx"  # Path to file to be processed by VMD
        name = "Spectral Data/Original spectral data/TL_Data1/Combined_gas_data_vmd.xlsx"  # Generate VMD processing file path
        num = 4  # The value of K in the VMD, decomposed into K eigencomponents
        del_files(name)  # Delete cache files
        data_x, data_y = read_data(path)  # Read the file
        u, u_hat, omega, uu_ = run_VMD(data_x, num)  # Execute VMD
        writeinexcel(uu_, data_y, name, num)  # The results are written to an excel file
    else:
        path = "Spectral Data/Original spectral data/TL_Data2/Mixed_gas_data.xlsx"  # Path to file to be processed by VMD
        name = "Spectral Data/Original spectral data/TL_Data2/Mixed_gas_data_vmd.xlsx"  # Generate VMD processing file path
        num = 4  # The value of K in the VMD, decomposed into K eigencomponents
        del_files(name)  # Delete cache files
        data_x, data_y = read_data(path)  # Read the file
        u, u_hat, omega, uu_ = run_VMD(data_x, num)  # Execute VMD
        writeinexcel(uu_, data_y, name, num)  # The results are written to an excel file
