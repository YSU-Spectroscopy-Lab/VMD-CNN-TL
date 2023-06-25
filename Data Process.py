import os
import numpy as np
import xlwt
import re
import pandas as pd
from openpyxl import Workbook
import shutil


def read_data(path, newpath, txt_num):
    '''
    * read_Data.py file is placed under the same root directory as the read data folder.
    * path：Enter the read data folder path.
    * Read the data folder layout as shown in the example.
    * After rerunning to read the data, if you rerun to read, you need to delete the newly generated **_ok folder in the Data folder before starting the operation.
    '''
    os.mkdir(newpath)
    path = path
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        root_ = []
        dirs_ = []
        a = 0
        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        file_name_after = newpath + '\\' + file_name_list + '_ok'
        if not os.path.exists(file_name_after):
            os.mkdir(file_name_after)
        for i in root_[1:]:
            file_name_after_ = file_name_after + '\\' + dirs__[a]
            if not os.path.exists(file_name_after_):
                os.mkdir(file_name_after_)
            txt_ = []
            for file_name_ in os.listdir(i):
                txt_.append(i + '\\' + file_name_)
            txt_num = txt_num  # 取多少文件
            txt_ = txt_[-txt_num:]
            num_all = []
            for txt_name in txt_:
                contents = []
                with open(txt_name, 'r') as f:
                    for line in f.readlines():
                        line = line.split('\n')
                        line = line[0].split('\t')
                        line = list(map(float, line))
                        contents.append(line)
                for content in contents:
                    num_all.append("%.4f" % (float(content[1]) / txt_num))
                if len(num_all) > len(contents):
                    for ii in range(len(num_all)):
                        if ii < len(contents):
                            num_all[ii] = "%.4f" % (num_all[ii] + float(num_all[ii + len(contents)]))
                        else:
                            num_all.pop()
                num_all = list(map(float, num_all))
                f.close()
                txt_name_after = newpath + '\\' + file_name_list + '_ok\\' + dirs__[a] + "\\" + dirs__[a] + ".txt"
                with open(txt_name_after, "w") as ff:
                    for li in num_all:
                        ff.write(str(li) + "\n")
                ff.close()
            a += 1
        print(file_name_list, "Data reading completed！")
    print("All data read completed！")


# Deducting backbone
def remove_bd(newpath):
    path = newpath
    nn_ = []
    file_name_lists = []
    for file_name in os.listdir(path):
        file_name_lists.append(file_name)
    for file_name_list in file_name_lists:
        n_p = path + '\\' + file_name_list + '_z'
        nn_.append(n_p)
        os.mkdir(n_p)
        root_ = []
        dirs_ = []

        for root, dirs, files in os.walk(path + '\\' + file_name_list):
            root_.append(root)
            dirs_.append(dirs)
            dirs__ = dirs_[0]
        root_.pop(0)
        root__ = root_[-1]
        dirs___ = dirs__[-1]
        root_.pop()
        dirs__.pop()
        bd_name = root__ + '\\' + dirs___ + '.txt'
        for i in range(len(root_)):
            data = []
            file_name = root_[i] + '\\' + dirs__[i] + '.txt'
            file_name_ = n_p + '\\' + dirs__[i] + '.txt'
            with open(bd_name) as bd_f:
                bd_file = bd_f.read().split('\n')
            bd_f.close()
            with open(file_name, 'r+') as f:
                file = f.read().split('\n')
            f.close()
            del (bd_file[-1])
            del (file[-1])
            bd_file = list(map(float, bd_file))
            file = list(map(float, file))
            for i in range(len(bd_file)):
                i_num = "%.4f" % ((file[i]) / (bd_file[i]))
                data.append(i_num)
            with open(file_name_, 'w') as f_:
                f_.truncate(0)
                for ii in data:
                    f_.write(ii + "\n")
            f_.close()
    print('Successfully removed the backing!')
    return nn_


# Gas mixtures written to excel
def writeinexcel_mixed(path):
    lu = []
    path = path
    le_ = 0
    wb1 = xlwt.Workbook(encoding='utf-8')  # Create a new excel file
    w1 = wb1.add_sheet('one')  # Add a new table with the name first
    ipath_ = path[0]
    file_name_lists_ = []
    file_name_lists = []
    for file_name_ in os.listdir(ipath_):
        file_name_ = re.findall('\d+', file_name_)
        b = file_name_[0] + ' ' + file_name_[1]
        ans = "".join(map(str, b))
        file_name_lists_.append(ans)
    file_name_lists_.sort(key=lambda x: int(re.findall('\d+', x)[0]))
    le = len(file_name_lists_)
    for le_i in range(le):
        w1.write(0, le_i + le_, file_name_lists_[le_i])
    le_ = le_ + len(file_name_lists_)
    for file_name_ in os.listdir(ipath_):
        file_name_lists.append(file_name_)

    file_name_lists.sort(key=lambda x: int(re.findall('\d+', x)[0]))
    # print("file_name_lists:",file_name_lists)
    for i_a in range(len(file_name_lists)):
        path_ = ipath_ + '\\' + file_name_lists[i_a]
        ii = 1
        for line in open(path_, encoding='utf-8'):
            if line == '\n':
                continue
            else:
                w1.write(ii, i_a, float(line))
                ii += 1

    wb1.save(path[0] + ".xls")
    p1 = path[0] + ".xls"
    lu.append(p1)

    if len(path) > 1:
        ipath__ = path[1]
        le__ = 0
        lei = 0
        wb2 = xlwt.Workbook(encoding='utf-8')  # Create a new excel file
        w2 = wb2.add_sheet('one')  # Add a new table with the name first
        file_name_lists_ = []
        file_name_lists = []
        for file_name_ in os.listdir(ipath__):
            file_name_ = re.sub('\D', '', file_name_)
            b = list(file_name_)
            b.pop()
            ans = "".join(map(str, b))
            file_name_lists_.append(ans)
        file_name_lists_ = list(map(int, file_name_lists_))
        file_name_lists_.sort()
        lei = len(file_name_lists_)
        for le_i in range(lei):
            w2.write(0, le_i + le__, file_name_lists_[le_i])
        le__ = le__ + len(file_name_lists_)
        for file_name_ in os.listdir(ipath__):
            file_name_lists.append(file_name_)
        file_name_lists.sort(key=lambda x: int(x[:-10]))
        for i_a in range(len(file_name_lists)):
            path_ = ipath__ + '\\' + file_name_lists[i_a]
            ii = 1
            for line in open(path_, encoding='utf-8'):
                if line == '\n':
                    continue
                else:
                    w2.write(ii, i_a, float(line))
                    ii += 1
                    # print(line)
        wb2.save(path[1] + ".xls")
        p2 = path[1] + ".xls"
        lu.append(p2)

    return lu


# Standard single gas data written to excel
def writeinexcel(path, nn):
    lu = []
    path = path
    le_ = 0
    wb1 = xlwt.Workbook(encoding='utf-8')
    w1 = wb1.add_sheet('one')
    ipath_ = path[0]
    file_name_lists_ = []
    file_name_lists = []
    for file_name_ in os.listdir(ipath_):
        file_name_ = re.sub('\D', '', file_name_)
        b = list(file_name_)
        b.pop()
        ans = "".join(map(str, b))
        file_name_lists_.append(ans)
    file_name_lists_ = list(map(int, file_name_lists_))
    file_name_lists_.sort()
    le = len(file_name_lists_)
    for le_i in range(le):
        w1.write(0, le_i + le_, file_name_lists_[le_i])
    le_ = le_ + len(file_name_lists_)
    for file_name_ in os.listdir(ipath_):
        file_name_lists.append(file_name_)
    file_name_lists.sort(key=lambda x: int(x[:-nn]))
    for i_a in range(len(file_name_lists)):
        path_ = ipath_ + '\\' + file_name_lists[i_a]
        ii = 1
        for line in open(path_, encoding='utf-8'):
            if line == '\n':
                continue
            else:
                w1.write(ii, i_a, float(line))
                ii += 1
    wb1.save(path[0] + ".xls")
    p1 = path[0] + ".xls"
    lu.append(p1)

    if len(path) > 1:
        ipath__ = path[1]
        le__ = 0
        wb2 = xlwt.Workbook(encoding='utf-8')
        w2 = wb2.add_sheet('one')
        file_name_lists_ = []
        file_name_lists = []
        for file_name_ in os.listdir(ipath__):
            file_name_ = re.sub('\D', '', file_name_)
            b = list(file_name_)
            b.pop()
            ans = "".join(map(str, b))
            file_name_lists_.append(ans)
        file_name_lists_ = list(map(int, file_name_lists_))
        file_name_lists_.sort()
        lei = len(file_name_lists_)
        for le_i in range(lei):
            w2.write(0, le_i + le__, file_name_lists_[le_i])
        le__ = le__ + len(file_name_lists_)
        for file_name_ in os.listdir(ipath__):
            file_name_lists.append(file_name_)
        file_name_lists.sort(key=lambda x: int(x[:-nn]))
        for i_a in range(len(file_name_lists)):
            path_ = ipath__ + '\\' + file_name_lists[i_a]
            ii = 1
            for line in open(path_, encoding='utf-8'):
                if line == '\n':
                    continue
                else:
                    w2.write(ii, i_a, float(line))
                    ii += 1
        wb2.save(path[1] + ".xls")
        p2 = path[1] + ".xls"
        lu.append(p2)

    return lu


# Reading data from excel
def readfromexcel(lu, add_data=False):
    wb3 = Workbook()
    wb3.create_sheet(index=0, title="all")
    w3 = wb3.active
    path1 = lu[0]
    path2 = lu[1]
    data1 = pd.read_excel(path1)
    columns1 = data1.columns
    data2 = pd.read_excel(path2)
    columns2 = data2.columns
    le = 0
    if add_data:
        for i in range(data1.shape[1]):
            for j in range(data2.shape[1]):
                lie1 = data1[columns1[i]]
                lie2 = data2[columns2[j]]
                '''Combination rules'''
                data_1 = lie1 + lie2
                w3.cell(1, j + le + 1, str(columns1[i]) + ' ' + str(columns2[j]))
                for ii in range(len(data_1)):
                    w3.cell(ii + 1 + 1, j + le + 1, data_1[ii])
            le += int(data2.shape[1])

        for time in range(40):
            a = "%.3f" % (np.random.ranf())
            b = "%.3f" % (np.random.ranf())
            if a == 0 or b == 0:
                a += 0.1
                b += 0.1
            for i in range(data1.shape[1]):
                for j in range(data2.shape[1]):
                    lie1 = data1[columns1[i]]
                    lie2 = data2[columns2[j]]
                    '''Combination rules'''
                    a = float(a)
                    b = float(b)
                    data_1 = a * lie1 + b * lie2

                    columns1_i = "%.4f" % (a * float(columns1[i]))
                    columns2_j = "%.4f" % (b * float(columns2[j]))
                    w3.cell(1, j + le + 1, str(columns1_i) + ' ' + str(columns2_j))
                    for ii in range(len(data_1)):
                        w3.cell(ii + 1 + 1, j + le + 1, data_1[ii])
                le += int(data2.shape[1])
            print("Circulation", time, "time")
    else:
        for i in range(data1.shape[1]):
            for j in range(data2.shape[1]):
                lie1 = data1[columns1[i]]
                lie2 = data2[columns2[j]]
                '''Combination rules'''
                data_1 = lie1 + lie2
                w3.cell(1, j + le + 1, str(columns1[i]) + ' ' + str(columns2[j]))
                for ii in range(len(data_1)):
                    w3.cell(ii + 1 + 1, j + le + 1, data_1[ii])
            le += int(data2.shape[1])

    pa = newpath + "\\data.xlsx"
    wb3.save(pa)
    print("Signal combination succeeded!")
    return pa


# Spectral slices
def cutsigleexcel(pa, a, b, c, d, ok):
    if ok:
        luu = []
        wb4 = xlwt.Workbook(encoding='utf-8')
        w4 = wb4.add_sheet('all')
        wb5 = xlwt.Workbook(encoding='utf-8')
        w5 = wb5.add_sheet('all')
        lu = pa
        path1 = lu[0]
        path2 = lu[1]
        data1 = pd.read_excel(path1)
        data2 = pd.read_excel(path2)
        data_1 = data1.iloc[c:d, :]
        data_2 = data2.iloc[c:d, :]
        data1 = data1.iloc[a:b, :]
        columns1 = data1.columns
        data2 = data2.iloc[a:b, :]
        columns2 = data2.columns
        data_1 = np.array(data_1)
        data1 = np.array(data1)
        data1_all = np.concatenate((data1, data_1), axis=0)
        data1_all = pd.DataFrame(data1_all, columns=columns1)
        data_2 = np.array(data_2)
        data2 = np.array(data2)
        data2_all = np.concatenate((data2, data_2), axis=0)
        data2_all = pd.DataFrame(data2_all, columns=columns2)
        print(data1_all)
        print(data2_all)
        for i in range(data1_all.shape[1]):
            lie_1 = data1_all[columns1[i]]
            w4.write(0, i, str(columns1[i]))
            for j in range(len(lie_1)):
                w4.write(j + 1, i, lie_1[j])
        for i in range(data2_all.shape[1]):
            lie_2 = data2_all[columns2[i]]
            w5.write(0, i, str(columns2[i]))
            for j in range(len(lie_2)):
                w5.write(j + 1, i, lie_2[j])
        path_1 = path1 + '_o.xls'
        wb4.save(path_1)
        path_2 = path2 + '_o.xls'
        wb5.save(path_2)
        luu.append(path_1)
        luu.append(path_2)
        print("Single signal slicing succeeded!")
    else:
        luu = []
        wb4 = xlwt.Workbook(encoding='utf-8')
        w4 = wb4.add_sheet('all')
        wb5 = xlwt.Workbook(encoding='utf-8')
        w5 = wb5.add_sheet('all')
        lu = pa
        path1 = lu[0]
        path2 = lu[1]
        data1 = pd.read_excel(path1)
        data2 = pd.read_excel(path2)
        data1 = data1.iloc[a:b, :]
        columns1 = data1.columns
        data2 = data2.iloc[a:b, :]
        columns2 = data2.columns
        for i in range(data1.shape[1]):
            lie_1 = data1[columns1[i]]
            w4.write(0, i, str(columns1[i]))
            for j in range(len(lie_1)):
                w4.write(j + 1, i, lie_1[a + j])
        for i in range(data2.shape[1]):
            lie_2 = data2[columns2[i]]
            w5.write(0, i, str(columns2[i]))
            for j in range(len(lie_2)):
                w5.write(j + 1, i, lie_2[a + j])
        path_1 = path1 + '_o.xls'
        wb4.save(path_1)
        path_2 = path2 + '_o.xls'
        wb5.save(path_2)
        luu.append(path_1)
        luu.append(path_2)
        print("Single signal slicing succeeded!")
    return luu


# Differential operation on standard single gases
def part_dif(lu, newpath):
    p = []
    wb6 = xlwt.Workbook(encoding='utf-8')
    w6 = wb6.add_sheet('all')
    wb7 = xlwt.Workbook(encoding='utf-8')
    w7 = wb7.add_sheet('all')
    path1 = lu[0]
    path2 = lu[1]
    data1 = pd.read_excel(path1)
    data2 = pd.read_excel(path2)
    columns1 = data1.columns
    columns2 = data2.columns
    data1 = data1.T
    data1 = np.array(data1)
    data1_deal = koumanbian(data1)
    data1_deal = pd.DataFrame(data1_deal)
    data1_deal = data1_deal.T
    data2 = data2.T
    data2 = np.array(data2)
    data2_deal = koumanbian(data2)
    data2_deal = pd.DataFrame(data2_deal)
    data2_deal = data2_deal.T
    columns1_ = data1_deal.columns
    columns2_ = data2_deal.columns

    for i in range(data1_deal.shape[1]):
        lie_1 = data1_deal[columns1_[i]]
        w6.write(0, i, str(columns1[i]))
        for j in range(len(lie_1)):
            w6.write(j + 1, i, lie_1[j])
    for i in range(data2_deal.shape[1]):
        lie_2 = data2_deal[columns2_[i]]
        w7.write(0, i, str(columns2[i]))
        for j in range(len(lie_2)):
            w7.write(j + 1, i, lie_2[j])
    path_1 = newpath + '\\data_1.xls'
    wb6.save(path_1)
    path_2 = newpath + '\\data_2.xls'
    wb7.save(path_2)
    p.append(path_1)
    p.append(path_2)
    print("Signal differential successful!")
    print("Please change the concentration to the true concentration！！！！！！！！！！！！！！！！！！！！！！！！！！")
    return p


# Mixed gas differential operation
def difall(pa, a, b, c, d, name, ok):
    if ok:
        name = name
        wb = Workbook()
        wb.create_sheet(index=0, title="all")
        ws = wb.active
        path = pa
        data = pd.read_excel(path)
        data1 = data.iloc[a:b, :]
        data2 = data.iloc[c:d, :]
        print("Signal slicing succeeded!")
        columns = data.columns
        data1 = np.array(data1)
        data2 = np.array(data2)
        data_all = np.concatenate((data1, data2), axis=0)
        data_all = pd.DataFrame(data_all, columns=columns)

        shape_ = data_all.shape[0]
        data_all = data_all.T
        data_all = np.array(data_all)
        data_deal = koumanbian(data_all)
        print("Signal differential successful!")
        data_deal = pd.DataFrame(data_deal)
        columns_ = data_deal.columns
        l = len(columns_)
        for i in range(shape_):
            lie = data_deal[columns_[i]]
            for j in range(len(lie)):
                ws.cell(j + 1, i + 1, lie[j])
        for col_i in range(len(columns)):
            col = str(columns[col_i]).split(' ')
            ws.cell(col_i + 1, l + 1, int(col[0]))
            ws.cell(col_i + 1, l + 2, int(col[1]))
        wb.save(name)
        print("Dataset creation completed!")
    else:
        name = name
        wb = Workbook()
        wb.create_sheet(index=0, title="all")
        ws = wb.active
        path = pa
        data = pd.read_excel(path)
        data = data.iloc[a:b, :]
        print("Signal slicing succeeded!")
        columns = data.columns
        shape_ = data.shape[0]
        data = data.T
        data = np.array(data)
        data_deal = koumanbian(data)
        print("Signal differential successful!")
        data_deal = pd.DataFrame(data_deal)
        columns_ = data_deal.columns
        l = len(columns_)
        for i in range(shape_):
            lie = data_deal[columns_[i]]
            for j in range(len(lie)):
                ws.cell(j + 1, i + 1, lie[j])
        for col_i in range(len(columns)):
            col = str(columns[col_i]).split(' ')
            ws.cell(col_i + 1, l + 1, int(col[0]))
            ws.cell(col_i + 1, l + 2, int(col[1]))
        wb.save(name)
        print("Dataset creation completed!")


# dif
def koumanbian(x):
    aa = []
    for i in x:
        zz1 = np.polyfit([i for i in range(len(i))], i, 10)
        pp1 = np.poly1d(zz1)
        aa.append(np.log(i / (pp1([i for i in range(len(i))]))))
    return np.array(aa)


# Building the dataset
def build_Data(pa, name):
    name = name
    wb = Workbook()
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    path = pa
    data = pd.read_excel(path)
    columns = data.columns
    shape_ = data.shape[0]
    data = data.T
    columns_ = data.columns
    l = len(columns_)
    for i in range(shape_):
        lie = data[columns_[i]]
        for j in range(len(lie)):
            ws.cell(j + 1, i + 1, lie[j])
    for col_i in range(len(columns)):
        col = str(columns[col_i]).split(' ')
        if len(col[0].split('.'))>1:
            col_0 = col[0].split('.')
            col[0] = col_0[0] + '.' + col_0[1]
        if len(col[1].split('.')) > 1:
            col_1 = col[1].split('.')
            col[1] = col_1[0] + '.' + col_1[1]
        ws.cell(col_i + 1, l + 1, float(col[0]))
        ws.cell(col_i + 1, l + 2, float(col[1]))
    wb.save(name)
    print("Dataset creation completed!")


# Clean up the data environment
def del_files(path, name):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    if os.path.exists(name):
        shutil.rmtree(name, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


# Differential fitting of mixed gases
def com_dif(lu, a, b, c, d, name, ok, dir):
    os.makedirs(dir)
    difall(lu[0], a - 1, b, c - 1, d, name, ok)  # Slicing, fitting and exporting the integrated data


# Standard single gas first differential then fit
def dif_com_first(lu, a, b, c, d, name, ok, dir,isadd = False):
    os.makedirs(dir)

    lu_ = cutsigleexcel(lu, a - 1, b, c - 1, d, ok)  # Slicing and dicing data to suit your needs

    pa_ = part_dif(lu_, newpath)  # After fitting the sliced data separately

    return pa_


# Standard single gas first differential then fit
def dif_com(name,isadd = False):

    pa_ = ['Spectral Data/Original spectral data/Standard single gases_/data_1.xls','Spectral Data/Original spectral data/Standard single gases_/data_2.xls']
    pa = readfromexcel(pa_, isadd)  # Integration of the two fitted data, with or without amplified data

    build_Data(pa, name)  # Create the data set and generate the data_end.xlsx file in the root directory

if __name__ == '__main__':
    '''
    When running is_single_gas = True:
        When prompted to change the concentration, please modify the data_1.xls and data_2.xls 
        in the Spectral Data/Original spectral data/Standard single gases_ directory.
        @@@@@@@@@@@@@@ The true concentrations are provided in Spectral Data/Original spectral data . 
        Then change is_real_c to True.
        Set is_real_c to False after running.
        The run time depends on the computer configuration and may be a few minutes.
    
    When running is_single_gas = False:
        After the program has been executed, modify the concentration in the 
        Spectral Data/Original spectral data/TL_Data2/Mixed_gas_data.xlsx file to the true concentration.
        @@@@@@@@@@@@@@@@ The true concentrations are in Spectral Data/Original spectral data/True concentrations.exel and
         Spectral Data/Original spectral data updated/True concentrations.exel files.
        
        Attention:
        After processing the data in the Spectral Data/Original spectral data/Mixed gases folder has been completed, 
        as the experimental data were collected on a different backing
        To process the data within the Spectral Data/Original spectral data updated/Mixed gases folder, 
        simply change the path parameters for oldpath, newpath, dir, and end_file_name.
        Finally, paste the contents of the generated Mixed_gas_data.xlsx into 
        the Spectral Data/Original spectral data/TL_Data2/Mixed_gas_data.xlsx file obtained in the previous step, 
        resulting in Spectral Data/Original spectral data/TL_Data2/Mixed_gas_data.xlsx file with 63 rows and 425 columns.
    '''
    is_single_gas = False
    is_real_c = False
    if is_single_gas:

        isadd = True  # To expand or not to expand the data set
        oldpath = 'Spectral Data/Original spectral data/Standard single gases'  # Original Data Location
        newpath = "Spectral Data/Original spectral data/Standard single gases_"  # Generate Process Data Location
        dir = "Spectral Data/Original spectral data/TL_Data1"  # Final generation of dataset catalog
        end_file_name = 'Spectral Data/Original spectral data/TL_Data1/Combined_gas_data.xlsx'  # Final generation of dataset catalog
        txt_num = 30  # Average the last number of txt files
        nn = 10  # Number position in name
        a = 198  # Start position of required data
        b = 620  # End position of required data
        c = 1398  # The second band start value, which is not used here
        d = 1864  # The second band termination value, which is not used here
        ok = False  # Read data from two bands is ok?, The study did not use two bands

        if is_real_c:
            dif_com(end_file_name, isadd=isadd)  # Standard single gas first differential then fit
        else:
            del_files(newpath, dir)  # Clean up the data environment

            read_data(oldpath, newpath, txt_num)  # Read the original data, average it, and store it in a new directory

            path1 = remove_bd(newpath)  # Remove data from the back

            path2 = writeinexcel(path1, nn)  # Standard single gas saved as excel

            pa_ = dif_com_first(path2, a, b, c, d, end_file_name, ok, dir, isadd= isadd)  # Standard single gas first differential then fit

    else:
        oldpath = 'Spectral Data/Original spectral data/Mixed gases'  # Original Data Location
        newpath = "Spectral Data/Original spectral data/Mixed gases_"  # Generate Process Data Location
        dir = "Spectral Data/Original spectral data/TL_Data2"  # Final generation of dataset catalog
        end_file_name = 'Spectral Data/Original spectral data/TL_Data2/Mixed_gas_data.xlsx'  # Final generation of dataset catalog
        # oldpath = 'Spectral Data/Original spectral data updated/Mixed gases'  # Original Data Location
        # newpath = "Spectral Data/Original spectral data updated/Mixed gases_"  # Generate Process Data Location
        # dir = "Spectral Data/Original spectral data updated/TL_Data2"  # Final generation of dataset catalog
        # end_file_name = 'Spectral Data/Original spectral data updated/TL_Data2/Mixed_gas_data.xlsx'  # Final generation of dataset catalog
        txt_num = 30  # Average the last number of txt files
        a = 198  # Start position of required data
        b = 620  # End position of required data
        c = 1398  # The second band start value, which is not used here
        d = 1864  # The second band termination value, which is not used here
        ok = False  # Read data from two bands is ok?, The study did not use two bands

        del_files(newpath, dir)  # Clean up the data environment

        read_data(oldpath, newpath, txt_num)  # Read the original data, average it, and store it in a new directory

        path1 = remove_bd(newpath)  # Remove data from the back

        path2 = writeinexcel_mixed(path1)  # Gas mixture data saved as excel

        com_dif(path2, a, b, c, d, end_file_name, ok, dir)  # Differential fitting of mixed gases
        print("Please change the concentration to the true concentration！！！！！！！！！！！！！！！！！！！！！！！！！！")

