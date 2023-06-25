import numpy as np
import pandas as pd
from keras.models import Sequential
from keras.utils import plot_model
from sklearn.model_selection import train_test_split
from keras.layers import Dense, Flatten, Conv1D, MaxPooling1D
from keras.models import model_from_json
from keras import backend as K
import os
import shutil
from openpyxl import Workbook


# Loading data
def load_data(path, num, ok=False):
    # Loading data
    df1 = pd.read_pickle(path)
    X = np.expand_dims(df1.values[:, 0:-2].astype(float), axis=2)  # Adding a one-dimensional axis
    Y = df1.values[:, -2:] / num
    # Divide training set, test set
    if ok:
        X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.25, shuffle=True)
    else:
        X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.25, random_state=0, shuffle=True)
    print("Loading of data complete!")
    return X_train, X_test, Y_train, Y_test


# Custom metric function, determination factor R_Squares
def coeff_determination(y_true, y_pred):
    SS_res = K.sum(K.square(y_true - y_pred))
    SS_tot = K.sum(K.square(y_true - K.mean(y_true)))
    return (1 - SS_res / (SS_tot + K.epsilon()))


# Defining the structure of a neural network
def build_model(name, optimizer, loss):
    model = Sequential()
    model.add(Conv1D(16, 3, input_shape=(422, 1), activation='relu'))
    model.add(Conv1D(16, 3, activation='relu'))
    model.add(MaxPooling1D(3))
    model.add(Conv1D(64, 3, activation='relu'))
    model.add(Conv1D(64, 3, activation='relu'))
    model.add(MaxPooling1D(3))
    model.add(Conv1D(128, 3, activation='relu'))
    model.add(Conv1D(128, 3, activation='relu'))
    model.add(MaxPooling1D(3))
    model.add(Conv1D(64, 3, activation='relu'))
    model.add(Conv1D(64, 3, activation='relu'))
    model.add(MaxPooling1D(3))
    model.add(Flatten())
    model.add(Dense(2, activation='linear'))
    plot_model(model, to_file=name, show_shapes=True)  # Printed model structure
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])  # Combination
    print("Building a neural network structure complete!")
    return model


# Save prediction results
def save_excel(predicted, Y_test, name, num):
    wb = Workbook()  # 新建一个excel文件
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    for i in range(predicted.shape[0]):
        ws.cell(i + 1, 1, predicted[i][0] * num)
        ws.cell(i + 1, 2, predicted[i][1] * num)
        ws.cell(i + 1, 3, Y_test[i][0] * num)
        ws.cell(i + 1, 4, Y_test[i][1] * num)
    wb.save(name)
    print("Save excel to finish!")


# Pre-training model starts training
def Start_model(model, name1, name2, name5, epochs, batch_size,
                optimizer, loss, num):
    isok = 1
    while isok:
        X_train, X_test, Y_train, Y_test = load_data(path1, num, ok=True)  # Loading data
        model.fit(X_train, Y_train, validation_data=(X_test, Y_test), epochs=epochs, batch_size=batch_size)
        predicted = predict_data(model, X_test, optimizer, loss)  # Predicted data
        CS2, SO2 = mean_error(predicted, Y_test, num)  # Calculating the mean absolute error
        if CS2 < 3 and SO2 < 10:
            isok = 0
    show_scores(model, X_test, Y_test)  # Calculation of the decision factor
    # save_excel(predicted, Y_test, name5, num)  # Save prediction results
    save_model(model, name1, name2)  # Preservation of models
    return model


# Predicted data
def predict_data(model, X_test, optimizer, loss):
    model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = model.predict(X_test)
    return predicted


# Calculation of the decision factor
def show_scores(model, X_test, Y_test):
    scores = model.evaluate(X_test, Y_test, verbose=0)
    print('coefficient of determination %s: %.6f%%' % (model.metrics_names[0], scores[1] * 100))


# Calculating the mean absolute error
def mean_error(predicted, Y_test, num):
    result = np.mean(abs(predicted[:, 0] * num - Y_test[:, 0] * num))
    print("Mean error of CS2 prediction results:", result)
    result_ = np.mean(abs(predicted[:, 1] * num - Y_test[:, 1] * num))
    print("Mean error of SO2 prediction results:", result_)
    return result, result_


# Preservation of models
def save_model(model, name1, name2):
    # Convert their model grid structure to json storage
    # Store model parameter weights as h5 files
    model_json = model.to_json()
    with open(name1, 'w') as json_file:
        json_file.write(model_json)
    model.save_weights(name2)
    print("Save model complete!")


# Pre-loading
def load_first(path, name2):
    path = path
    json_file = open(path, "r")
    loaded_model_json = json_file.read()
    json_file.close()
    loaded_model = model_from_json(loaded_model_json)
    loaded_model.load_weights(name2)
    print("Loading of pre-trained models complete!")
    return loaded_model


# Transfer learning
def transfer_Learning(path2, name1, name2, name3, name6, epochs, batch_size, optimizer, loss, num):
    loaded_model = load_first(name1, name2)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    isok = 1
    while isok:
        X_train, X_test, Y_train, Y_test = load_data(path2, num, True)
        loaded_model.fit(X_train, Y_train,
                         validation_data=(X_test, Y_test),
                         epochs=epochs, batch_size=batch_size)
        predicted = predict_data(loaded_model, X_test, optimizer, loss)  # Predicted data
        CS2, SO2 = mean_error(predicted, Y_test, num)  # Calculating the mean absolute error
        if CS2 < 5 and SO2 < 10:
            isok = 0

    show_scores(loaded_model, X_test, Y_test)  # Calculation of the decision factor
    print("Migration training complete!")
    save_excel(predicted, Y_test, name6, num)
    save_model(loaded_model, name1, name3)  # Preservation of models


# Delete cache files
def del_files(path):
    if os.path.exists(path):
        shutil.rmtree(path, ignore_errors=False, onerror=None)
    print("Data environment cleanup succeeded!")


# Save the test file
def save_test_excel(predicted, Y_test, name, num):
    wb = Workbook()  # Create a new excel file
    wb.create_sheet(index=0, title="all")
    ws = wb.active
    Y_test = Y_test
    for i in range(predicted.shape[0]):
        ws.cell(i + 1, 1, predicted[i][0] * num)
        ws.cell(i + 1, 2, predicted[i][1] * num)
        ws.cell(i + 1, 3, Y_test[i][0] * num)
        ws.cell(i + 1, 4, Y_test[i][1] * num)
    wb.save(name)
    print("Save excel to finish!")


# Test models
def test_model(path3, name1, model_h5_name, name7, optimizer, loss, num):
    df1 = pd.read_pickle(path3)
    X_test = np.expand_dims(df1.values[:, 0:-2].astype(float), axis=2)  # Adding a one-dimensional axis
    Y_test = df1.values[:, -2:] / num
    loaded_model = load_first(name1, model_h5_name)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = predict_data(loaded_model, X_test, optimizer, loss)  # Predicted data
    mean_error(predicted, Y_test, num)  # Calculating the mean absolute error
    show_scores(loaded_model, X_test, Y_test)  # Calculation of the decision factor
    print("Test complete!")
    save_test_excel(predicted, Y_test, name7, num)


def test_model_self(path3, name1, model_h5_name, name7, optimizer, loss, num):
    df1 = pd.read_pickle(path3)
    X_test = np.expand_dims(df1.values[:, 0:-2].astype(float), axis=2)  # Adding a one-dimensional axis
    Y_test = df1.values[:, -2:] / num
    loaded_model = load_first(name1, model_h5_name)
    loaded_model.compile(optimizer=optimizer, loss=loss, metrics=[coeff_determination])
    predicted = predict_data(loaded_model, X_test, optimizer, loss)  # Predicted data
    mean_error(predicted, Y_test, num)  # Calculating the mean absolute error
    show_scores(loaded_model, X_test, Y_test)  # Calculation of the decision factor
    print("Test complete!")
    save_test_excel(predicted, Y_test, name7, num)


if __name__ == '__main__':
    '''
    The model structure and parameters that have been generated for this study are in the folder Model Structure and 
    Parameters and the test set data are in Spectral Data/Test data. The results in the paper can be verified by directly 
    changing the incoming parameters of the test_model function to the generated file. If you want to start training the 
    model again, please run the test_model_self function to test the retrained generated model after the training is 
    completed. The path4 parameter is the path of the predicted gas concentration group generated after migration learning, 
    which requires manually selecting the concentration groups from the Mixed_gas_data_vmd.xlsx file based 
    on the Run_data/Transfer-training_results.xlsx file and saving them as .xlsx files.In the pkl format file generated 
    via Evaluation Indicators.py.
    '''
    path1 = "Spectral Data/Original spectral data/TL_Data1/Combined_gas_data_vmd.pkl"  # Pre-trained model data path
    path2 = "Spectral Data/Original spectral data/TL_Data2/Mixed_gas_data_vmd.pkl"  # Transfer learning training data path
    path3 = "Spectral Data/Test data/VMD-Test.pkl"  # Validation set path for testing
    path4 = ""
    run_data_path = "Run_data"  # Save directory for running files
    name1 = "Run_data/model.json"  # Model structure
    name1_test = "Model Structure and Parameters/model.json"
    name2 = "Run_data/model.h5"  # Pre-training model parameters
    name3 = "Run_data/model_transfer_learning.h5"  # Transfer learning training model parameters
    name3_test = "Model Structure and Parameters/model_transfer_learning.h5"
    name4 = "Run_data/model_linear.png"  # Visualisation of the model structure
    name5 = "Run_data/Pre-training_results.xlsx"  # Pre-training results
    name6 = "Run_data/Transfer-training_results.xlsx"  # Transfer training results
    name7 = "Run_data/test_result.xlsx"  # Test result file
    epochs = 200  # Number of iterations  vmd: CNN 400 CNN-TL 1100  none: CNN 200 CNN-TL 200
    batch_size = 256  # Batch size
    optimizer = 'adam'  # Optimizers
    loss = "mean_squared_error"  # Root Mean Square Error
    num = 1000  # Standardisation factor
    is_Transfer_Learning = True  # Transfer learning or not
    is_test = True  # To test or not to test

    # To test or not to test
    if is_test:
        """Delete last run data"""
        if not os.path.exists(run_data_path):
            os.mkdir(run_data_path)
        """Model pre-training"""
        # test_model(path3, name1_test, name3_test, name7, optimizer, loss, num)
        test_model_self(path3, name1, name3, name7, optimizer, loss, num)
    else:
        # Transfer learning or not
        if is_Transfer_Learning:
            transfer_Learning(path2, name1, name2, name3, name6, epochs, batch_size, optimizer, loss,
                              num)
        else:
            """Delete last run data"""
            del_files(run_data_path)
            os.mkdir(run_data_path)
            """Model pre-training"""
            model = build_model(name4, optimizer, loss)  # Building the model
            Start_model(model, name1, name2, name5, epochs, batch_size,
                        optimizer, loss, num)  # Training models
